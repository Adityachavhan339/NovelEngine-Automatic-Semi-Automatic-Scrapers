import pandas as pd
import os
import shutil
import requests
from bs4 import BeautifulSoup
import time
import random
import openpyxl
from openpyxl.styles import PatternFill,Font
from openpyxl.formatting.rule import CellIsRule

def my_branding():
    print("NOVELENGINE PRO: STAGE 1 FINAL DEPLOYMENT")
    print("Creator: Aditya/Daniel")
    print("Github: Adityachavhan339")
    print("Please Leave Us Your Feedback On Github")
    

def pull_web_novels(selected_genre,num_pages):
    basic_url = f"https://freewebnovels.com/{selected_genre}/"
    stealth_headers = {"User-Agent": "Mozilla/5.0"}
    total_pulled_novels = []

    print(f"Starting Novel Extraction For {selected_genre}")

    for current_num in range(1,num_pages+1):
        try:
            selected_url = basic_url if current_num == 1 else f"{basic_url}{current_num}"
            print(f"Scraping Page NO: {current_num}")

            website_response = requests.get(basic_url,headers=stealth_headers)
            soup_Novels = BeautifulSoup(website_response.text,"lxml")
            raw_novels = soup_Novels.find_all("div",class_="li")


            for novel in raw_novels:
                name_tag = novel.find(class_="tit")
                rating_tag = novel.find(class_="core")
                chapter_count = novel.find("span",class_="s1")

                name = name_tag.get_text(strip=True)if name_tag else "N/A"
                rating = rating_tag.get_text(strip=True)if rating_tag else "N/A"
                ch_count = chapter_count.get_text(strip=True)if chapter_count else "N/A"

                total_pulled_novels.append({"Name": name,"Rating": rating,"Total Chapters": ch_count})

            time.sleep(random.uniform(1,3))
        except Exception as e:
            print(e)

    pulled_web_novelss = pd.DataFrame(total_pulled_novels)
    return(pulled_web_novelss)

def pro_way(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
    work_sheet = workbook.active

    work_sheet.freeze_panes = "A2"
    bold_font = Font(bold=True)

    for cell in work_sheet:
        cell.font = bold_font

    green_fill = PatternFill(start_color="00FF00",fill_type="solid")
    red_fill = PatternFill(start_color="FF0000",fill_type="solid")

    work_sheet.conditional_formatting.add('B2:B1000',CellIsRule(operator='greaterThan',formula=[4.5],fill=green_fill))
    work_sheet.conditional_formatting.add('B2:B1000',CellIsRule(operator='lessThan',formula=[2.5],fill=red_fill))

    work_sheet.conditional_formatting.add('C2:C1000',CellIsRule(operator='greaterThan',formula=[500],fill=green_fill))
    work_sheet.conditional_formatting.add('C2:C1000',CellIsRule(operator='lessThan',formula=[100],fill=red_fill))

    workbook.save(f"Super{excel_path}")
    print(f"Super Styled Applied To:{excel_path}")


def finishing_files(genre_type,data_main,data_rate,data_chapters):
    
    os.makedirs("Backup Of Day",exist_ok=True)
    raw_novel_csv = (f"Perfected_Novels_{genre_type}_Data.csv")

    data_main.to_csv(raw_novel_csv,index=False)

    shutil.copy(raw_novel_csv,"Backup Of Day")

    file_rating = (f"Rating_Novels_{genre_type}.xlsx")
    file_chapter = (f"Chapter_Novels_{genre_type}.xlsx")
    data_rate.to_excel(file_rating,index=False)
    data_chapters.to_excel(file_chapter,index=False)

    pro_way(file_rating)
    pro_way(file_chapter)

    final_folder = input("Enter Name For Pro Way Folder:")
    os.makedirs(final_folder)

    for file in [file_rating,file_chapter,f"PRO_WAY_{file_rating}",f"PRO_WAY{file_chapter}"]:
        shutil.move(file,os.path.join(final_folder,file))

    print(f"All Files Transferred At: {final_folder}")

def main():
    print(----"System Starting----")

    select_genre = input("Select Genre eg:Horror,Fantasy,Action:")
    select_pages = int(input("Select No. Of Pages [1 Page = 20 Novels]:"))

    master_key = pull_web_novels(select_genre,select_pages)
    print(f"Total Collected Novels:{len(master_key)}")

    master_key["Rating"] = master_key["Rating"].fillna("Not Rated")
    clean_master_data = master_key.dropna(subset=["Name"])

    sort_rating = clean_master_data.sort_values(by="Rating",ascending=False)
    sort_chapters = clean_master_data.sort_values(by="Total Chapters",ascending=False)

    finishing_files(select_genre,clean_master_data,sort_rating,sort_chapters)


    my_branding()

if __name__ == "__main__":
    main()