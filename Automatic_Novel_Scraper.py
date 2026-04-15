import pandas as pd
import shutil
import os
import requests
from bs4 import BeautifulSoup
import time
import random
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule

print("Starting Data Analysis")
genre = "Fantasy"
genre_url = f"https://freewebnovel.com/genre/{genre}/"
headers = {'User-Agent': 'Mozilla/5.0'}

novel_data = []
print("Starting Data Extraction..")
print("Collecting Data..")
for page_num in range(1,12):
    try:  
       url = genre_url if page_num == 1 else (f"{genre_url}{page_num}")

       print(f"Scrapping Page {page_num}..")
       response = requests.get(url,headers=headers)
       # Fixed typo 'soap' to 'soup' to match your next lines
       soup = BeautifulSoup(response.text,"lxml")

       raw_novel_data = soup.find_all("div",class_="li")

       for novel in raw_novel_data:
              name = "N/A"
              rating = "N/A"
              chapter = "N/A"

              name_tag = novel.find(class_="tit")
              rating_tag = novel.find(class_="core")
              chapter_tag = novel.find("span",class_="s1")

              if name_tag:
                   name = name_tag.get_text(strip=True)
              if rating_tag:
                   rating = rating_tag.get_text(strip=True)
              if chapter_tag:
                   chapter = chapter_tag.get_text(strip=True)

              novel_data.append({"Name": name,"Rating": rating,"chapter_count": chapter})
       time.sleep(random.uniform(1,4))

    except Exception as e:
          print(e)

print(f"Total Collected Novels {len(novel_data)}")
print("Starting Final Data Organization..")
updated_novel_data = pd.DataFrame(novel_data)

updated_novel_data.to_csv(f"Perfected_Novels_{genre}_Data.csv",index=False)

print("Starting Folder Creation..")
try:
   os.makedirs(f"{genre} Novels",exist_ok=True)
   print(f"Successfully Created Folder {genre} Novels")
except Exception as e:
     print(e)

print("Saving Files..")
print("Moving Files To Safe Folder")
try:
     os.makedirs("Backup_Of_Day",exist_ok=True)
     shutil.copy(f"Perfected_Novels_{genre}_Data.csv","Backup_Of_Day")
     shutil.move(f"Perfected_Novels_{genre}_Data.csv",os.path.join(f"{genre} Novels",f"Perfected_Novels_{genre}_Data.csv"))
     print(f"Operation Successful Saved CSV File At {genre} Novels")
except Exception as e:
     print(e)

data = pd.read_csv(f"{genre} Novels/Perfected_Novels_{genre}_Data.csv")
data["Rating"] = data["Rating"].fillna("Not Rated")
cleaned_data = data.dropna(subset=["Name"])

perfected_data_rating = cleaned_data.sort_values(by=["Rating"],ascending=False)
perfected_data_chapters = cleaned_data.sort_values(by=["chapter_count"],ascending=False)

try:
    os.makedirs(f"Sorted {genre} Novels", exist_ok=True)
    print("Folder Creation Successful")
except Exception as e:
    print(e)


perfected_data_rating.to_csv(f"Perfected_Rating_{genre}.csv",index=False)
perfected_data_chapters.to_csv(f"Perfected_Chapters_{genre}.csv",index=False)


file_ch = f"Perfected_Novels_{genre}_Chapter_Data.xlsx"
file_rt = f"Perfected_Novels_{genre}_Rating_Data.xlsx"
perfected_data_chapters.to_excel(file_ch, index=False)
perfected_data_rating.to_excel(file_rt, index=False)


for file in [file_ch, file_rt]:
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    ws.freeze_panes = "A2"
    header_font = Font(bold=True)
    for cell in ws[1]: # Bold the top row
        cell.font = header_font

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    ws.conditional_formatting.add('B2:B1000', CellIsRule(operator='greaterThan', formula=['4.5'], fill=green_fill))
    ws.conditional_formatting.add('B2:B1000', CellIsRule(operator='lessThan', formula=['2.5'], fill=red_fill))
    ws.conditional_formatting.add('C2:C1000', CellIsRule(operator='greaterThan', formula=['500'], fill=green_fill))
    ws.conditional_formatting.add('C2:C1000', CellIsRule(operator='lessThan', formula=['100'], fill=red_fill))
    wb.save(f"PRO_{file}")


try:
    shutil.move(f"Perfected_Rating_{genre}.csv", os.path.join(f"Sorted {genre} Novels", f"Perfected_Rating_{genre}.csv"))
    shutil.move(f"Perfected_Chapters_{genre}.csv", os.path.join(f"Sorted {genre} Novels", f"Perfected_Chapters_{genre}.csv"))
    shutil.move(file_ch, os.path.join(f"Sorted {genre} Novels", file_ch))
    shutil.move(file_rt, os.path.join(f"Sorted {genre} Novels", file_rt))
    shutil.move(f"PRO_{file_ch}", os.path.join(f"Sorted {genre} Novels", f"PRO_{file_ch}"))
    shutil.move(f"PRO_{file_rt}", os.path.join(f"Sorted {genre} Novels", f"PRO_{file_rt}"))
    print(f"Operation Successful Saved All Files At Sorted {genre} Novels")
except Exception as e:
    print(e)

print("Thank You For Using Our Service Please Leave Us Your Feedback At Github")
        
